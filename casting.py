import json
import time
import re
import os
from datetime import datetime
import google.generativeai as genai
from PIL import Image
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import fitz  # PyMuPDF for PDF processing

from casting_config import (
    MATERIAL_PROPERTIES, CASTING_TYPES,
    get_material_guidance, get_volume_guidance, get_recommended_action,
    get_filename_components
)

# Try to load from .env file if available
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# Configure Gemini API
api_key = os.getenv("GEMINI_API_KEY")
if not api_key:
    print("Warning: GEMINI_API_KEY not found. Please set your API key.")
    print("You can set it as an environment variable or in a .env file")
    exit(1)

genai.configure(api_key=api_key)

GEMINI_MODEL = "gemini-2.5-flash"
API_DELAY = 0.5
RULES_PATH = "input/rules.json"
OUTPUT_DIR = "output"

os.makedirs(OUTPUT_DIR, exist_ok=True)

model = genai.GenerativeModel(
    model_name=GEMINI_MODEL,
    generation_config={
        "temperature": 0,
        "response_mime_type": "application/json"
    }
)

def load_rules_from_json():
    """Load rules from the constant rules.json file"""
    with open(RULES_PATH, 'r', encoding='utf-8') as f:
        rules_data = json.load(f)
    
    rules = []
    for rule_data in rules_data:
        rule = {
            'rule_id': f"R{rule_data['rule_number']}",
            'title': rule_data['rule_header'],
            'engineering_intent': rule_data['hint_description'],
            'ai_guidance': rule_data['hint_description'],
            'checklist_items': []
        }
        
        for i, check_item in enumerate(rule_data['check_items'], 1):
            rule['checklist_items'].append({
                'check_id': f"{rule_data['rule_number']}.{i}",
                'text': check_item
            })
        
        rules.append(rule)
    
    return {"rules": rules}

def extract_json_from_response(response_text):
    """Extract JSON from AI response"""
    text = response_text.strip()
    
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    
    # Try extracting from markdown code blocks
    code_block_patterns = [
        r'```json\s*([\s\S]*?)\s*```',  
        r'```\s*([\s\S]*?)\s*```',     
    ]
    for pattern in code_block_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            try:
                return json.loads(match.group(1).strip())
            except json.JSONDecodeError:
                continue
    
    # Try finding JSON by brace matching
    brace_start = text.find('{')
    if brace_start != -1:
        depth = 0
        for i, char in enumerate(text[brace_start:], start=brace_start):
            if char == '{':
                depth += 1
            elif char == '}':
                depth -= 1
                if depth == 0:
                    try:
                        return json.loads(text[brace_start:i+1])
                    except json.JSONDecodeError:
                        break
    
    return None

def evaluate_checklist_item(rule, check_item, image, context):
    """Evaluate a single checklist item with the uploaded image"""
    prompt = f"""You are a senior casting design engineer. Analyze the casting drawing for a specific checklist item.

CASTING SPECIFICATIONS:
- Type: {context['casting_type']}
- Material: {context['material']}
- Production Volume: {context['volume']:,} parts
- Process: {context['process']}
- Tolerance: {context['tolerance']}
- Surface Finish: {context['surface_finish']}

RULE CONTEXT:
- Rule ID: {rule['rule_id']}
- Title: {rule['title']}
- Engineering Intent: {rule['engineering_intent']}

CHECKLIST ITEM TO EVALUATE:
- Check ID: {check_item['check_id']}
- Requirement: {check_item['text']}

MATERIAL-SPECIFIC CONSIDERATIONS:
{get_material_guidance(context['material'])}

VOLUME-SPECIFIC CONSIDERATIONS:
{get_volume_guidance(context['volume'])}

INSTRUCTIONS:
1. Examine the provided drawing carefully
2. Evaluate ONLY this specific checklist item based on VISIBLE geometry
3. Consider the material properties and production volume in your assessment
4. Factor in the casting type and process requirements

OUTPUT REQUIREMENTS - CRITICAL:
- Return ONLY a single JSON object
- Use EXACTLY this format:

{{"result": "Yes", "reason": "Brief engineering justification considering {context['material']} and {context['volume']:,} parts", "confidence": "High"}}

RESULT VALUES:
- "Yes" = Drawing complies with this checklist item
- "No" = Drawing violates this checklist item (explain the violation)
- "Needs Review" = Cannot determine from visible geometry

CONFIDENCE VALUES:
- "High" = Clear evidence visible in drawing
- "Medium" = Partial evidence, some assumptions made
- "Low" = Limited visibility, uncertain assessment

RESPOND WITH JSON ONLY:"""

    try:
        response = model.generate_content([prompt, image])
        response_text = response.text
        
        parsed = extract_json_from_response(response_text)
        
        if parsed:
            result = parsed.get("result", "Needs Review")
            reason = parsed.get("reason", "No reason provided")
            confidence = parsed.get("confidence", "Low")
            
            # Normalize result values
            result_upper = result.strip().lower()
            if result_upper in ["yes", "compliant", "pass", "true"]:
                result = "Yes"
            elif result_upper in ["no", "non-compliant", "fail", "false"]:
                result = "No"
            else:
                result = "Needs Review"
            
            return {
                "result": result,
                "reason": str(reason)[:500],
                "confidence": confidence
            }
        else:
            return {
                "result": "Needs Review",
                "reason": "AI response parsing failed",
                "confidence": "Low"
            }
            
    except Exception as e:
        return {
            "result": "Needs Review",
            "reason": f"API error: {str(e)}",
            "confidence": "Low"
        }

def analyze_casting_image(image_path, casting_context):
    """Main analysis function - simplified for single image input"""
    
    # Load rules from JSON
    rules_data = load_rules_from_json()
    
    # Load image
    image = Image.open(image_path)
    
    checklist_rows = []
    success_count = 0
    total_checks = 0

    for rule in rules_data["rules"]:
        print(f"Evaluating {rule['rule_id']} – {rule['title']}")

        for check_item in rule['checklist_items']:
            print(f"  Checking {check_item['check_id']}: {check_item['text'][:50]}...")
            
            total_checks += 1
            result = evaluate_checklist_item(rule, check_item, image, casting_context)
            
            if "parsing failed" not in result["reason"].lower() and "api error" not in result["reason"].lower():
                success_count += 1
            
            recommended_action = get_recommended_action(rule, check_item, result['result'], casting_context)
            
            checklist_rows.append({
                "Rule ID": rule["rule_id"],
                "Rule Title": rule["title"],
                "Check ID": check_item["check_id"],
                "Checklist Item": check_item["text"],
                "Result (Yes/No)": result["result"],
                "Notes / Observations": result["reason"],
                "Recommended Actions": recommended_action
            })
            
            time.sleep(API_DELAY)

    # Generate Excel report
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    material_short, volume_short = get_filename_components(casting_context)
    
    output_file = f"casting_analysis_{material_short}_{volume_short}parts_{timestamp}.xlsx"
    output_path = os.path.join(OUTPUT_DIR, output_file)

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Casting Analysis"
    
    # Add headers and formatting
    ws.merge_cells('A1:G1')
    ws['A1'] = f"CASTING DESIGN ANALYSIS - {casting_context['casting_type']}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Material: {casting_context['material']} | Volume: {casting_context['volume']:,} parts | Process: {casting_context['process']}"
    ws['A2'].font = Font(size=10)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:G3')
    ws['A3'] = f"Tolerance: {casting_context['tolerance']} | Surface Finish: {casting_context['surface_finish']} | Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A3'].font = Font(size=9)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Rule ID", "Rule Title", "Check ID", "Checklist Item", "Result (Yes/No)", "Notes / Observations", "Recommended Actions"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Data rows
    df = pd.DataFrame(checklist_rows)
    for r_idx, row in enumerate(df.values, 6):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            # Color code results
            if c_idx == 5:  # Result column
                if value == "Yes":
                    cell.fill = green_fill
                elif value == "No":
                    cell.fill = red_fill
    
    # Merge cells for Rule ID and Rule Title
    current_rule = None
    merge_start = 6
    
    for row_idx in range(6, len(df) + 7):
        if row_idx <= len(df) + 5:
            rule_id = ws.cell(row=row_idx, column=1).value
        else:
            rule_id = None
            
        if rule_id != current_rule:
            if current_rule is not None and row_idx - 1 > merge_start:
                ws.merge_cells(start_row=merge_start, start_column=1, end_row=row_idx-1, end_column=1)
                ws.cell(row=merge_start, column=1).alignment = Alignment(horizontal='center', vertical='center')
                
                ws.merge_cells(start_row=merge_start, start_column=2, end_row=row_idx-1, end_column=2)
                ws.cell(row=merge_start, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            merge_start = row_idx
            current_rule = rule_id
    
    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 60
    ws.column_dimensions['G'].width = 70
    
    wb.save(output_path)

    # Calculate summary statistics
    yes_count = sum(1 for row in checklist_rows if row["Result (Yes/No)"] == "Yes")
    no_count = sum(1 for row in checklist_rows if row["Result (Yes/No)"] == "No")
    review_count = sum(1 for row in checklist_rows if row["Result (Yes/No)"] == "Needs Review")

    print(f"\n=== ANALYSIS COMPLETE ===")
    print(f"Report saved to: {output_path}")
    print(f"Successfully evaluated: {success_count}/{total_checks} checklist items")
    print(f"✓ Compliant: {yes_count}")
    print(f"✗ Non-compliant: {no_count}")
    print(f"? Needs Review: {review_count}")

    return {
        "total_checks": total_checks,
        "successful_evaluations": success_count,
        "results": {
            "compliant": yes_count,
            "non_compliant": no_count,
            "needs_review": review_count
        },
        "output_file": output_file,
        "details": checklist_rows[:10]  # First 10 items for preview
    }

def get_user_inputs():
    """Get casting specifications and image path from user"""
    print("=== CASTING DESIGN ANALYSIS ===\n")
    
    # Get image/PDF path
    image_path = input("Enter path to casting drawing (PDF/PNG/JPG): ").strip()
    if not os.path.exists(image_path):
        print(f"Error: File not found: {image_path}")
        return None
    
    print("\nCasting Type Options:")
    casting_types = [
        "Sand Casting", "Shell Molding", "Investment Casting", "Lost Foam Casting",
        "Gravity Die Casting", "Low Pressure Die Casting", 
        "High Pressure Die Casting (Hot Chamber)", "High Pressure Die Casting (Cold Chamber)",
        "Centrifugal Casting", "Squeeze Casting"
    ]
    
    for i, casting_type in enumerate(casting_types, 1):
        print(f"{i}. {casting_type}")
    
    print(f"{len(casting_types) + 1}. Other (specify)")
    
    choice = input(f"\nSelect casting type (1-{len(casting_types) + 1}): ").strip()
    
    try:
        choice_num = int(choice)
        if 1 <= choice_num <= len(casting_types):
            casting_type = casting_types[choice_num - 1]
        elif choice_num == len(casting_types) + 1:
            casting_type = input("Enter custom casting type: ").strip()
        else:
            print("Invalid choice, using 'Sand Casting'")
            casting_type = "Sand Casting"
    except ValueError:
        print("Invalid input, using 'Sand Casting'")
        casting_type = "Sand Casting"
    
    print(f"\nMaterial Options:")
    materials = list(MATERIAL_PROPERTIES.keys())
    for i, material in enumerate(materials, 1):
        print(f"{i}. {material}")
    
    print(f"{len(materials) + 1}. Other (specify)")
    
    mat_choice = input(f"\nSelect material (1-{len(materials) + 1}): ").strip()
    
    try:
        mat_num = int(mat_choice)
        if 1 <= mat_num <= len(materials):
            material = materials[mat_num - 1]
        elif mat_num == len(materials) + 1:
            material = input("Enter custom material: ").strip()
        else:
            print("Invalid choice, using 'Gray Cast Iron'")
            material = "Gray Cast Iron"
    except ValueError:
        print("Invalid input, using 'Gray Cast Iron'")
        material = "Gray Cast Iron"
    
    try:
        volume = int(input("\nEnter production volume (number of parts): ").strip())
    except ValueError:
        print("Invalid input, using default volume of 100")
        volume = 100
    
    print(f"\nCasting Process Options:")
    processes = ["Sand Casting", "Investment Casting", "Die Casting", "Permanent Mold Casting", "Other"]
    for i, process in enumerate(processes, 1):
        print(f"{i}. {process}")
    
    process_choice = input(f"\nSelect casting process (1-{len(processes)}): ").strip()
    
    try:
        process_num = int(process_choice)
        if 1 <= process_num <= len(processes) - 1:
            suggested_process = processes[process_num - 1]
        elif process_num == len(processes):
            suggested_process = input("Enter custom casting process: ").strip()
        else:
            print("Invalid choice, using 'Sand Casting'")
            suggested_process = "Sand Casting"
    except ValueError:
        print("Invalid input, using 'Sand Casting'")
        suggested_process = "Sand Casting"
    
    tolerance = input("\nTolerance requirements (e.g., ±0.5mm) [Optional]: ").strip() or "Standard"
    surface_finish = input("Surface finish requirements (e.g., Ra 3.2) [Optional]: ").strip() or "As-cast"
    
    casting_context = {
        "casting_type": casting_type,
        "material": material,
        "volume": volume,
        "process": suggested_process,
        "tolerance": tolerance,
        "surface_finish": surface_finish
    }
    
    # Display summary
    print(f"\n=== ANALYSIS PARAMETERS ===")
    print(f"Drawing: {os.path.basename(image_path)}")
    print(f"Casting Type: {casting_type}")
    print(f"Material: {material}")
    print(f"Production Volume: {volume:,} parts")
    print(f"Casting Process: {suggested_process}")
    print(f"Tolerance: {tolerance}")
    print(f"Surface Finish: {surface_finish}")
    
    confirm = input(f"\nProceed with analysis? (y/n): ").strip().lower()
    if confirm != 'y':
        print("Analysis cancelled.")
        return None
    
    return image_path, casting_context

def convert_pdf_to_image(pdf_path):
    """Convert PDF to image if needed"""
    if pdf_path.lower().endswith('.pdf'):
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(pdf_path)
            page = doc.load_page(0)  # Get first page
            
            zoom = 300 / 72  # 300 DPI
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            
            # Save as temporary image
            temp_image_path = pdf_path.replace('.pdf', '_temp.png')
            pix.save(temp_image_path)
            doc.close()
            
            print(f"Converted PDF to image: {temp_image_path}")
            return temp_image_path
        except Exception as e:
            print(f"Error converting PDF: {e}")
            return None
    else:
        return pdf_path  # Already an image

if __name__ == "__main__":
    user_inputs = get_user_inputs()
    if user_inputs is None:
        exit()
    
    file_path, casting_context = user_inputs
    
    print(f"\n=== STARTING ANALYSIS ===")
    
    # Convert PDF to image if needed
    image_path = convert_pdf_to_image(file_path)
    if image_path is None:
        print("Error: Could not process the input file.")
        exit()
    
    try:
        # Run the analysis
        result = analyze_casting_image(image_path, casting_context)
        
        print(f"\n=== ANALYSIS COMPLETE ===")
        print(f"Excel report generated: {result['output_file']}")
        print(f"Total checks: {result['total_checks']}")
        print(f"Successful evaluations: {result['successful_evaluations']}")
        print(f"Results: {result['results']['compliant']} compliant, {result['results']['non_compliant']} non-compliant, {result['results']['needs_review']} needs review")
        
        # Clean up temporary image if created
        if image_path != file_path and os.path.exists(image_path):
            os.remove(image_path)
            
    except Exception as e:
        print(f"Error during analysis: {e}")
        # Clean up temporary image if created
        if image_path != file_path and os.path.exists(image_path):
            os.remove(image_path)
