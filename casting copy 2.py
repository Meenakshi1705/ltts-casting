import json
import time
import re
# import vertexai
# from vertexai.generative_models import GenerativeModel, Part
import google.generativeai as genai

import pandas as pd
from datetime import datetime
import os
# from pdf2image import convert_from_path
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from casting_config import (
    MATERIAL_PROPERTIES, CASTING_TYPES,
    get_material_guidance, get_volume_guidance, get_recommended_action,
    get_process_suggestion, get_filename_components
)

# PROJECT_ID = "gemini-gdandt-01"
# GEMINI_LOCATION = "us-central1"

# Try to load from .env file if available
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

api_key = os.getenv("GEMINI_API_KEY")
if not api_key:
    print("Warning: GEMINI_API_KEY not found. Please set your API key.")
    print("You can set it as an environment variable or in a .env file")
    exit(1)

genai.configure(api_key=api_key)
GEMINI_MODEL = "gemini-2.5-flash"
API_DELAY = 2
API_TIMEOUT = 120  

RULES_PATH = r"D:\casting\input\Casting_Design_Checklist.xlsx"
OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# vertexai.init(project=PROJECT_ID, location=GEMINI_LOCATION)
# model = GenerativeModel(GEMINI_MODEL)

model = genai.GenerativeModel(
    model_name=GEMINI_MODEL,
    generation_config={
        "temperature": 0,
        "response_mime_type": "application/json"
    }
)

def load_rules_from_excel(excel_path):
    df = pd.read_excel(excel_path)
    
    rules = []
    current_rule = None
    
    for _, row in df.iterrows():
        rule_number = row['#']
        rule_header = row['Rule / Header']
        hint_description = row['Hint / Description']
        check_item = row['Check Item']
        
        if pd.notna(rule_number):
            if current_rule is not None:
                rules.append(current_rule)
            
            current_rule = {
                'rule_id': f"R{int(rule_number)}",
                'title': rule_header,
                'engineering_intent': hint_description if pd.notna(hint_description) else "",
                'ai_guidance': hint_description if pd.notna(hint_description) else "",
                'checklist_items': []
            }
        
        if pd.notna(check_item) and current_rule is not None:
            rule_num = int(current_rule['rule_id'][1:])
            item_count = len(current_rule['checklist_items']) + 1
            check_id = f"{rule_num}.{item_count}"
            
            current_rule['checklist_items'].append({
                'check_id': check_id,
                'text': check_item
            })
    
    if current_rule is not None:
        rules.append(current_rule)
    
    return {"rules": rules}

    pages = convert_from_path(pdf_path, dpi=300)
    image_paths = []

    for i, page in enumerate(pages):
        img_path = os.path.join(output_dir, f"page_{i+1}.png")
        page.save(img_path, "PNG")
        image_paths.append(img_path)

    return image_paths


def pdf_to_images(pdf_path, output_dir="images", dpi=300):

    os.makedirs(output_dir, exist_ok=True)

    doc = fitz.open(pdf_path)
    image_paths = []

    zoom = dpi / 72  # 72 DPI is PDF base
    mat = fitz.Matrix(zoom, zoom)

    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        pix = page.get_pixmap(matrix=mat, alpha=False)

        img_path = os.path.join(output_dir, f"page_{page_num + 1}.png")
        pix.save(img_path)
        image_paths.append(img_path)

    doc.close()
    return image_paths


def extract_json_from_response(response_text):
    """
    Robust JSON extraction that handles:
    - Markdown code blocks (```json ... ```)
    - Extra text before/after JSON
    - Multiple JSON objects (takes first valid one)
    """
    text = response_text.strip()
    
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    
    code_block_patterns = [
        r'```json\s*([\s\S]*?)\s*```',  
        r'```\s*([\s\S]*?)\s*```',     
    ]
    for pattern in code_block_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            try:
                return json.loads(match.group(1).strip())
            except json.JSONDecodeError:
                continue
    
    brace_start = text.find('{')
    if brace_start != -1:
        depth = 0
        for i, char in enumerate(text[brace_start:], start=brace_start):
            if char == '{':
                depth += 1
            elif char == '}':
                depth -= 1
                if depth == 0:
                    try:
                        return json.loads(text[brace_start:i+1])
                    except json.JSONDecodeError:
                        break
    
    json_pattern = r'\{\s*"result"\s*:\s*"[^"]*"\s*,\s*"reason"\s*:\s*"[^"]*"\s*,\s*"confidence"\s*:\s*"[^"]*"\s*\}'
    match = re.search(json_pattern, text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(0))
        except json.JSONDecodeError:
            pass
    
    return None


def evaluate_checklist_item(rule, check_item, image_parts, context):
    """
    Evaluate a single checklist item with customized context
    """
    prompt = f"""You are a senior casting design engineer. Analyze the 2D casting drawing for a specific checklist item.

CASTING SPECIFICATIONS:
- Type: {context['casting_type']}
- Material: {context['material']}
- Production Volume: {context['volume']:,} parts
- Suggested Process: {context['process']}
- Tolerance: {context['tolerance']}
- Surface Finish: {context['surface_finish']}

RULE CONTEXT:
- Rule ID: {rule['rule_id']}
- Title: {rule['title']}
- Engineering Intent: {rule['engineering_intent']}
- Guidance: {rule['ai_guidance']}

CHECKLIST ITEM TO EVALUATE:
- Check ID: {check_item['check_id']}
- Requirement: {check_item['text']}

MATERIAL-SPECIFIC CONSIDERATIONS:
{get_material_guidance(context['material'])}

VOLUME-SPECIFIC CONSIDERATIONS:
{get_volume_guidance(context['volume'])}

INSTRUCTIONS:
1. Examine ALL provided drawing views carefully
2. Evaluate ONLY this specific checklist item based on VISIBLE geometry
3. Consider the material properties and production volume in your assessment
4. Factor in the casting type and process requirements

OUTPUT REQUIREMENTS - CRITICAL:
- Return ONLY a single JSON object
- NO markdown, NO code blocks, NO backticks, NO explanation text
- Use EXACTLY this format:

{{"result": "Yes", "reason": "Brief engineering justification considering {context['material']} and {context['volume']:,} parts", "confidence": "High"}}

RESULT VALUES:
- "Yes" = Drawing complies with this checklist item for the specified material and volume
- "No" = Drawing violates this checklist item (explain the violation)
- "Needs Review" = Cannot determine from visible geometry

CONFIDENCE VALUES:
- "High" = Clear evidence visible in drawing
- "Medium" = Partial evidence, some assumptions made
- "Low" = Limited visibility, uncertain assessment

RESPOND WITH JSON ONLY:"""

    try:
        response = model.generate_content([prompt] + image_parts)
        response_text = response.text
        
        parsed = extract_json_from_response(response_text)
        
        if parsed:
            result = parsed.get("result", "Needs Review")
            reason = parsed.get("reason", "No reason provided")
            confidence = parsed.get("confidence", "Low")
            
            result_upper = result.strip().lower()
            if result_upper in ["yes", "compliant", "pass", "true"]:
                result = "Yes"
            elif result_upper in ["no", "non-compliant", "fail", "false"]:
                result = "No"
            else:
                result = "Needs Review"
            
            # conf_upper = confidence.strip().lower()
            # if conf_upper in ["high", "h"]:
            #     confidence = "High"
            # elif conf_upper in ["medium", "med", "m", "moderate"]:
            #     confidence = "Medium"
            # else:
            #     confidence = "Low"
            
            return {
                "result": result,
                "reason": str(reason)[:500],
                "confidence": confidence
            }
        else:
            print(f"[{check_item['check_id']}] JSON extraction failed.")
            return {
                "result": "Needs Review",
                "reason": f"AI response parsing failed",
                "confidence": "Low"
            }
            
    except Exception as e:
        print(f"[{check_item['check_id']}] API error: {str(e)}")
        return {
            "result": "Needs Review",
            "reason": f"API error: {str(e)}",
            "confidence": "Low"
        }

def get_user_inputs():
    """Get casting specifications from user"""
    print("=== CASTING DESIGN ANALYSIS v1 ===\n")
    
    excel_path = input("Enter path to casting rules Excel file: ").strip()
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found: {excel_path}")
        return None
    
    pdf_path = input("Enter path to casting drawing PDF: ").strip()
    if not os.path.exists(pdf_path):
        print(f"Error: PDF file not found: {pdf_path}")
        return None
    
    print("\nCasting Type Options:")
    casting_types = [
        "Sand Casting",
        "Shell Molding",
        "Investment Casting",
        "Lost Foam Casting",
        "Gravity Die Casting",
        "Low Pressure Die Casting",
        "High Pressure Die Casting (Hot Chamber)",
        "High Pressure Die Casting (Cold Chamber)",
        "Centrifugal Casting",
        "Squeeze Casting"
    ]

    
    for i, casting_type in enumerate(casting_types, 1):
        print(f"{i}. {casting_type}")
    
    print(f"{len(casting_types) + 1}. Other (specify)")
    
    choice = input(f"\nSelect casting type (1-{len(casting_types) + 1}): ").strip()
    
    try:
        choice_num = int(choice)
        if 1 <= choice_num <= len(casting_types):
            casting_type = casting_types[choice_num - 1]
        elif choice_num == len(casting_types) + 1:
            casting_type = input("Enter custom casting type: ").strip()
        else:
            print("Invalid choice, using 'General Casting'")
            casting_type = "General Casting"
    except ValueError:
        print("Invalid input, using 'General Casting'")
        casting_type = "General Casting"
    
    print(f"\nMaterial Options:")
    materials = list(MATERIAL_PROPERTIES.keys())
    for i, material in enumerate(materials, 1):
        print(f"{i}. {material}")
    
    print(f"{len(materials) + 1}. Other (specify)")
    
    mat_choice = input(f"\nSelect material (1-{len(materials) + 1}): ").strip()
    
    try:
        mat_num = int(mat_choice)
        if 1 <= mat_num <= len(materials):
            material = materials[mat_num - 1]
        elif mat_num == len(materials) + 1:
            material = input("Enter custom material: ").strip()
        else:
            print("Invalid choice, using 'Gray Cast Iron'")
            material = "Gray Cast Iron"
    except ValueError:
        print("Invalid input, using 'Gray Cast Iron'")
        material = "Gray Cast Iron"
    
    try:
        volume = int(input("\nEnter production volume (number of parts): ").strip())
    except ValueError:
        print("Invalid input, using default volume of 100")
        volume = 100
    
    print(f"\nCasting Process Options:")
    processes = ["Sand Casting", "Investment Casting", "Die Casting", "Permanent Mold Casting", "Other"]
    for i, process in enumerate(processes, 1):
        print(f"{i}. {process}")
    
    process_choice = input(f"\nSelect casting process (1-{len(processes)}): ").strip()
    
    try:
        process_num = int(process_choice)
        if 1 <= process_num <= len(processes) - 1:
            suggested_process = processes[process_num - 1]
        elif process_num == len(processes):
            suggested_process = input("Enter custom casting process: ").strip()
        else:
            print("Invalid choice, using 'To Be Determined'")
            suggested_process = "To Be Determined"
    except ValueError:
        print("Invalid input, using 'To Be Determined'")
        suggested_process = "To Be Determined"
    
    tolerance = input("\nTolerance requirements (e.g., ±0.5mm) [Optional]: ").strip() or "Standard"
    surface_finish = input("Surface finish requirements (e.g., Ra 3.2) [Optional]: ").strip() or "As-cast"
    
    casting_context = {
        "casting_type": casting_type,
        "material": material,
        "volume": volume,
        "process": suggested_process,
        "tolerance": tolerance,
        "surface_finish": surface_finish
    }
    
    # Display summary
    print(f"\n=== ANALYSIS PARAMETERS ===")
    print(f"Rules Excel: {os.path.basename(excel_path)}")
    print(f"Drawing PDF: {os.path.basename(pdf_path)}")
    print(f"Casting Type: {casting_type}")
    print(f"Material: {material}")
    print(f"Production Volume: {volume:,} parts")
    print(f"Casting Process: {suggested_process}")
    print(f"Tolerance: {tolerance}")
    print(f"Surface Finish: {surface_finish}")
    
    confirm = input(f"\nProceed with analysis? (y/n): ").strip().lower()
    if confirm != 'y':
        print("Analysis cancelled.")
        return None
    
    return excel_path, pdf_path, casting_context

def evaluate_rule(rule, image_parts, context):
    """
    Evaluate all checklist items within a rule.
    Returns a list of results for each sub-item.
    """
    results = []
    
    for check_item in rule['checklist_items']:
        print(f"  Checking {check_item['check_id']}: {check_item['text'][:50]}...")
        
        result = evaluate_checklist_item(rule, check_item, image_parts, context)
        
        recommended_action = get_recommended_action(rule, check_item, result['result'], context)
        
        results.append({
            "check_id": check_item['check_id'],
            "check_text": check_item['text'],
            "result": result['result'],
            "reason": result['reason'],
            "recommended_action": recommended_action
        })
        
        time.sleep(API_DELAY)  
    
    return results

def run_rule_engine():
    """Main function with user input and customized analysis"""
    
    user_inputs = get_user_inputs()
    if user_inputs is None:
        return
    
    excel_path, pdf_path, casting_context = user_inputs
    
    print(f"\n=== STARTING ANALYSIS ===")
    
    print("Loading rules from Excel...")
    rules_data = load_rules_from_excel(excel_path)

    print("Converting PDF to images...")
    image_paths = pdf_to_images(pdf_path)
    
    print("Loading images for AI analysis...")
    image_parts = []
    for img in image_paths:
        from PIL import Image
        image = Image.open(img)
        image_parts.append(image)

    checklist_rows = []
    success_count = 0
    total_checks = 0

    for rule in rules_data["rules"]:
        print(f"\nEvaluating {rule['rule_id']} – {rule['title']}")

        sub_results = evaluate_rule(rule, image_parts, casting_context)
        
        for sub in sub_results:
            total_checks += 1
            if "parsing failed" not in sub["reason"].lower() and "api error" not in sub["reason"].lower():
                success_count += 1
            
            checklist_rows.append({
                "Rule ID": rule["rule_id"],
                "Rule Title": rule["title"],
                "Check ID": sub["check_id"],
                "Checklist Item": sub["check_text"],
                "Result (Yes/No)": sub["result"],
                "Notes / Observations": sub["reason"],
                # "Confidence": sub["confidence"],
                "Recommended Actions": sub["recommended_action"]
            })

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    material_short, volume_short = get_filename_components(casting_context)
    
    output_file = os.path.join(
        OUTPUT_DIR,
        f"casting_analysis_{material_short}_{volume_short}parts_{timestamp}.xlsx"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Casting Analysis"
    
    ws.merge_cells('A1:H1')
    ws['A1'] = f"CASTING DESIGN ANALYSIS - {casting_context['casting_type']}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:H2')
    ws['A2'] = f"Material: {casting_context['material']} | Volume: {casting_context['volume']:,} parts | Process: {casting_context['process']}"
    ws['A2'].font = Font(size=10)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:H3')
    ws['A3'] = f"Tolerance: {casting_context['tolerance']} | Surface Finish: {casting_context['surface_finish']} | Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A3'].font = Font(size=9)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ["Rule ID", "Rule Title", "Check ID", "Checklist Item", "Result (Yes/No)", "Notes / Observations", "Recommended Actions"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    df = pd.DataFrame(checklist_rows)
    for r_idx, row in enumerate(df.values, 6):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            if c_idx == 5:
                if value == "Yes":
                    cell.fill = green_fill
                elif value == "No":
                    cell.fill = red_fill
    
    current_rule = None
    merge_start = 6
    
    for row_idx in range(6, len(df) + 7):
        if row_idx <= len(df) + 5:
            rule_id = ws.cell(row=row_idx, column=1).value
        else:
            rule_id = None
            
        if rule_id != current_rule:
            if current_rule is not None and row_idx - 1 > merge_start:
                ws.merge_cells(start_row=merge_start, start_column=1, end_row=row_idx-1, end_column=1)
                ws.cell(row=merge_start, column=1).alignment = Alignment(horizontal='center', vertical='center')
                
                ws.merge_cells(start_row=merge_start, start_column=2, end_row=row_idx-1, end_column=2)
                ws.cell(row=merge_start, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            merge_start = row_idx
            current_rule = rule_id
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 60
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 70
    
    wb.save(output_file)

    print(f"\n=== ANALYSIS COMPLETE ===")
    print(f"Report saved to: {output_file}")
    print(f"Successfully evaluated: {success_count}/{total_checks} checklist items")
    
    yes_count = sum(1 for row in checklist_rows if row["Result (Yes/No)"] == "Yes")
    no_count = sum(1 for row in checklist_rows if row["Result (Yes/No)"] == "No")
    review_count = sum(1 for row in checklist_rows if row["Result (Yes/No)"] == "Needs Review")
    
    print(f"\nResults Summary:")
    print(f"✓ Compliant: {yes_count}")
    print(f"✗ Non-compliant: {no_count}")
    print(f"? Needs Review: {review_count}")
    
    if no_count > 0:
        print(f"\n⚠️  {no_count} issues found - check Recommended Actions column for solutions")
    
    print(f"\nAnalysis customized for:")
    print(f"- {casting_context['material']} casting")
    print(f"- {casting_context['volume']:,} parts production volume")
    print(f"- {casting_context['process']} process")


if __name__ == "__main__":
    run_rule_engine()
