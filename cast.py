import json
import time
import re
import vertexai
from vertexai.generative_models import GenerativeModel, Part
import pandas as pd
from datetime import datetime
import os
from pdf2image import convert_from_path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


PROJECT_ID = "gemini-gdandt-01"
GEMINI_LOCATION = "us-central1"
GEMINI_MODEL = "gemini-2.5-flash"
API_DELAY = 2
API_TIMEOUT = 120  

RULES_PATH = r"D:\casting\rules\rules.json"
PDF_PATH = r"D:\casting\input\Connector Casting Input.pdf"

OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

CASTING_CONTEXT = {
    "casting_type": "Bracket Casting",
    "material": "Gray Cast Iron",
    "process": "Sand Casting"
}

vertexai.init(project=PROJECT_ID, location=GEMINI_LOCATION)

model = GenerativeModel(GEMINI_MODEL)

def pdf_to_images(pdf_path, output_dir="images"):
    os.makedirs(output_dir, exist_ok=True)

    pages = convert_from_path(pdf_path, dpi=300)
    image_paths = []

    for i, page in enumerate(pages):
        img_path = os.path.join(output_dir, f"page_{i+1}.png")
        page.save(img_path, "PNG")
        image_paths.append(img_path)

    return image_paths


def extract_json_from_response(response_text):
    """
    Robust JSON extraction that handles:
    - Markdown code blocks (```json ... ```)
    - Extra text before/after JSON
    - Multiple JSON objects (takes first valid one)
    """
    text = response_text.strip()
    
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    
    code_block_patterns = [
        r'```json\s*([\s\S]*?)\s*```',  
        r'```\s*([\s\S]*?)\s*```',     
    ]
    for pattern in code_block_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            try:
                return json.loads(match.group(1).strip())
            except json.JSONDecodeError:
                continue
    
    brace_start = text.find('{')
    if brace_start != -1:
        depth = 0
        for i, char in enumerate(text[brace_start:], start=brace_start):
            if char == '{':
                depth += 1
            elif char == '}':
                depth -= 1
                if depth == 0:
                    try:
                        return json.loads(text[brace_start:i+1])
                    except json.JSONDecodeError:
                        break
    
        json_pattern = r'\{\s*"result"\s*:\s*"[^"]*"\s*,\s*"reason"\s*:\s*"[^"]*"\s*,\s*"confidence"\s*:\s*"[^"]*"\s*\}'
    match = re.search(json_pattern, text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(0))
        except json.JSONDecodeError:
            pass
    
    return None


def evaluate_checklist_item(rule, check_item, image_parts, context):
    """
    Evaluate a single checklist item within a rule.
    """
    prompt = f"""You are a senior casting design engineer. Analyze the 2D casting drawing for a specific checklist item.

CASTING CONTEXT:
- Type: {context['casting_type']}
- Material: {context['material']}
- Process: {context['process']}

RULE CONTEXT:
- Rule ID: {rule['rule_id']}
- Title: {rule['title']}
- Engineering Intent: {rule['engineering_intent']}
- Guidance: {rule['ai_guidance']}

CHECKLIST ITEM TO EVALUATE:
- Check ID: {check_item['check_id']}
- Requirement: {check_item['text']}

INSTRUCTIONS:
1. Examine ALL provided drawing views carefully
2. Evaluate ONLY this specific checklist item based on VISIBLE geometry
3. Consider the engineering intent and defect risks

OUTPUT REQUIREMENTS - CRITICAL:
- Return ONLY a single JSON object
- NO markdown, NO code blocks, NO backticks, NO explanation text
- NO text before or after the JSON
- Use EXACTLY this format:

{{"result": "Yes", "reason": "Brief engineering justification", "confidence": "High"}}

RESULT VALUES:
- "Yes" = Drawing complies with this checklist item
- "No" = Drawing violates this checklist item (explain the violation)
- "Needs Review" = Cannot determine from visible geometry

CONFIDENCE VALUES:
- "High" = Clear evidence visible in drawing
- "Medium" = Partial evidence, some assumptions made
- "Low" = Limited visibility, uncertain assessment

RESPOND WITH JSON ONLY:"""

    try:
        response = model.generate_content([prompt] + image_parts)
        response_text = response.text
        
        parsed = extract_json_from_response(response_text)
        
        if parsed:
            result = parsed.get("result", "Needs Review")
            reason = parsed.get("reason", "No reason provided")
            confidence = parsed.get("confidence", "Low")
            
            result_upper = result.strip().lower()
            if result_upper in ["yes", "compliant", "pass", "true"]:
                result = "Yes"
            elif result_upper in ["no", "non-compliant", "fail", "false"]:
                result = "No"
            else:
                result = "Needs Review"
            
            conf_upper = confidence.strip().lower()
            if conf_upper in ["high", "h"]:
                confidence = "High"
            elif conf_upper in ["medium", "med", "m", "moderate"]:
                confidence = "Medium"
            else:
                confidence = "Low"
            
            return {
                "result": result,
                "reason": str(reason)[:500],
                "confidence": confidence
            }
        else:
            print(f"[{check_item['check_id']}] JSON extraction failed.")
            return {
                "result": "Needs Review",
                "reason": f"AI response parsing failed",
                "confidence": "Low"
            }
            
    except Exception as e:
        print(f"[{check_item['check_id']}] API error: {str(e)}")
        return {
            "result": "Needs Review",
            "reason": f"API error: {str(e)}",
            "confidence": "Low"
        }


def evaluate_rule(rule, image_parts, context):
    """
    Evaluate all checklist items within a rule.
    Returns a list of results for each sub-item.
    """
    results = []
    
    for check_item in rule['checklist_items']:
        print(f"  Checking {check_item['check_id']}: {check_item['text'][:50]}...")
        
        result = evaluate_checklist_item(rule, check_item, image_parts, context)
        results.append({
            "check_id": check_item['check_id'],
            "check_text": check_item['text'],
            "result": result['result'],
            "reason": result['reason'],
            "confidence": result['confidence']
        })
        
        time.sleep(API_DELAY)  
    
    return results

def run_rule_engine():
    with open(RULES_PATH, "r") as f:
        rules_data = json.load(f)

    image_paths = pdf_to_images(PDF_PATH)
    
    image_parts = []
    for img in image_paths:
        with open(img, "rb") as f:
            image_parts.append(Part.from_data(data=f.read(), mime_type="image/png"))

    checklist_rows = []
    success_count = 0
    total_checks = 0

    for rule in rules_data["rules"]:
        print(f"\nEvaluating {rule['rule_id']} – {rule['title']}")

        sub_results = evaluate_rule(rule, image_parts, CASTING_CONTEXT)
        
        for sub in sub_results:
            total_checks += 1
            if "parsing failed" not in sub["reason"].lower() and "api error" not in sub["reason"].lower():
                success_count += 1
            
            checklist_rows.append({
                "Rule ID": rule["rule_id"],
                "Rule Title": rule["title"],
                "Check ID": sub["check_id"],
                "Checklist Item": sub["check_text"],
                "Result (Yes/No)": sub["result"],
                "Notes / Observations": sub["reason"],
                "Confidence": sub["confidence"]
            })


    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(
        OUTPUT_DIR,
        f"casting_checklist_{timestamp}.xlsx"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Casting Checklist"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ["Rule ID", "Rule Title", "Check ID", "Checklist Item", "Result (Yes/No)", "Notes / Observations", "Confidence"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    df = pd.DataFrame(checklist_rows)
    for r_idx, row in enumerate(df.values, 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            # Color code Result column (column 5)
            if c_idx == 5:
                if value == "Yes":
                    cell.fill = green_fill
                elif value == "No":
                    cell.fill = red_fill
    
    current_rule = None
    merge_start = 2
    
    for row_idx in range(2, len(df) + 3):
        if row_idx <= len(df) + 1:
            rule_id = ws.cell(row=row_idx, column=1).value
        else:
            rule_id = None 
            
        if rule_id != current_rule:
            if current_rule is not None and row_idx - 1 > merge_start:
                ws.merge_cells(start_row=merge_start, start_column=1, end_row=row_idx-1, end_column=1)
                ws.cell(row=merge_start, column=1).alignment = Alignment(horizontal='center', vertical='center')
                
                ws.merge_cells(start_row=merge_start, start_column=2, end_row=row_idx-1, end_column=2)
                ws.cell(row=merge_start, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            merge_start = row_idx
            current_rule = rule_id
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 60
    ws.column_dimensions['G'].width = 12
    
    wb.save(output_file)

    print(f"\nChecklist saved to {output_file}")
    print(f"Successfully evaluated: {success_count}/{total_checks} checklist items")


if __name__ == "__main__":
    run_rule_engine()
