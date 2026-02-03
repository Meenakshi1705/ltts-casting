import pandas as pd
import os
import sys
import json
import time
import re
from datetime import datetime
import fitz  # PyMuPDF
import google.generativeai as genai

# Add parent directory to path to import our modules
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    from casting_config import (
        MATERIAL_PROPERTIES, CASTING_TYPES,
        get_material_guidance, get_volume_guidance, get_recommended_action,
        get_process_suggestion, get_filename_components
    )
    
    # Try to load from .env file if available
    try:
        from dotenv import load_dotenv
        load_dotenv()
    except ImportError:
        pass
    
    # Temporarily disable API due to quota limits - use mock mode
    GEMINI_AVAILABLE = False
    print("Warning: Using mock mode due to API quota limits.")
except ImportError:
    GEMINI_AVAILABLE = False
    print("Warning: Gemini API not available. Running in mock mode.")

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Gemini API Configuration
GEMINI_MODEL = "gemini-1.5-flash"
API_DELAY = 0.5

if GEMINI_AVAILABLE:
    model = genai.GenerativeModel(
        model_name=GEMINI_MODEL,
        generation_config={
            "temperature": 0,
            "response_mime_type": "application/json"
        }
    )


def load_rules_from_excel(excel_path):
    """Load rules from Excel checklist file"""
    df = pd.read_excel(excel_path)
    
    rules = []
    current_rule = None
    
    for _, row in df.iterrows():
        rule_number = row['#']
        rule_header = row['Rule / Header']
        hint_description = row['Hint / Description']
        check_item = row['Check Item']
        
        # If we have a rule number, this is a new rule header
        if pd.notna(rule_number):
            if current_rule is not None:
                rules.append(current_rule)
            
            current_rule = {
                'rule_id': f"R{int(rule_number)}",
                'title': rule_header,
                'engineering_intent': hint_description if pd.notna(hint_description) else "",
                'ai_guidance': hint_description if pd.notna(hint_description) else "",
                'checklist_items': []
            }
        
        # If we have a check item, add it to the current rule
        if pd.notna(check_item) and current_rule is not None:
            rule_num = int(current_rule['rule_id'][1:])
            item_count = len(current_rule['checklist_items']) + 1
            check_id = f"{rule_num}.{item_count}"
            
            current_rule['checklist_items'].append({
                'check_id': check_id,
                'text': check_item
            })
    
    # Add the last rule
    if current_rule is not None:
        rules.append(current_rule)
    
    return {"rules": rules}


def pdf_to_images(pdf_path, output_dir="temp_images"):
    """Convert PDF to images using PyMuPDF"""
    os.makedirs(output_dir, exist_ok=True)
    
    doc = fitz.open(pdf_path)
    image_paths = []

    zoom = 300 / 72  # 300 DPI
    mat = fitz.Matrix(zoom, zoom)

    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        pix = page.get_pixmap(matrix=mat, alpha=False)

        img_path = os.path.join(output_dir, f"page_{page_num + 1}.png")
        pix.save(img_path)
        image_paths.append(img_path)

    doc.close()
    return image_paths


def extract_json_from_response(response_text):
    """Robust JSON extraction"""
    text = response_text.strip()
    
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    
    # Try extracting from markdown code blocks
    code_block_patterns = [
        r'```json\s*([\s\S]*?)\s*```',  
        r'```\s*([\s\S]*?)\s*```',     
    ]
    for pattern in code_block_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            try:
                return json.loads(match.group(1).strip())
            except json.JSONDecodeError:
                continue
    
    # Try finding JSON by brace matching
    brace_start = text.find('{')
    if brace_start != -1:
        depth = 0
        for i, char in enumerate(text[brace_start:], start=brace_start):
            if char == '{':
                depth += 1
            elif char == '}':
                depth -= 1
                if depth == 0:
                    try:
                        return json.loads(text[brace_start:i+1])
                    except json.JSONDecodeError:
                        break
    
    return None


def evaluate_checklist_item_mock(rule, check_item, context):
    """Enhanced mock evaluation with realistic data"""
    import random
    
    # More realistic distribution of results
    results = ["Yes"] * 6 + ["No"] * 2 + ["Needs Review"] * 2  # 60% Yes, 20% No, 20% Review
    result = random.choice(results)
    
    if result == "No":
        reasons = [
            f"Wall thickness below minimum {context['material']} requirement",
            f"Missing draft angles for {context['casting_type']} process",
            f"Sharp corners detected - add fillets for {context['material']}",
            f"Insufficient feeding paths for {context['volume']:,} parts production",
            f"Undercuts present - redesign for {context['process']} process"
        ]
        reason = random.choice(reasons)
    elif result == "Yes":
        reasons = [
            f"Compliant with {context['material']} casting requirements",
            f"Proper design for {context['casting_type']} process",
            f"Adequate wall thickness for {context['volume']:,} parts",
            f"Good draft angles for {context['process']} manufacturing",
            f"Appropriate geometry for {context['material']} properties"
        ]
        reason = random.choice(reasons)
    else:
        reasons = [
            f"Drawing lacks detail for {context['casting_type']} evaluation",
            f"Cannot determine compliance from visible geometry",
            f"Additional views needed for {context['material']} assessment",
            f"Insufficient information for {context['process']} analysis"
        ]
        reason = random.choice(reasons)
    
    return {
        "result": result,
        "reason": reason,
        "confidence": random.choice(["High", "Medium", "Low"])
    }


def evaluate_checklist_item(rule, check_item, image_parts, context):
    """Evaluate a single checklist item with customized context using direct Gemini API"""
    if not GEMINI_AVAILABLE:
        return evaluate_checklist_item_mock(rule, check_item, context)
    
    prompt = f"""You are a senior casting design engineer. Analyze the 2D casting drawing for a specific checklist item.

CASTING SPECIFICATIONS:
- Type: {context['casting_type']}
- Material: {context['material']}
- Production Volume: {context['volume']:,} parts
- Suggested Process: {context['process']}
- Tolerance: {context['tolerance']}
- Surface Finish: {context['surface_finish']}

RULE CONTEXT:
- Rule ID: {rule['rule_id']}
- Title: {rule['title']}
- Engineering Intent: {rule['engineering_intent']}
- Guidance: {rule['ai_guidance']}

CHECKLIST ITEM TO EVALUATE:
- Check ID: {check_item['check_id']}
- Requirement: {check_item['text']}

MATERIAL-SPECIFIC CONSIDERATIONS:
{get_material_guidance(context['material'])}

VOLUME-SPECIFIC CONSIDERATIONS:
{get_volume_guidance(context['volume'])}

INSTRUCTIONS:
1. Examine ALL provided drawing views carefully
2. Evaluate ONLY this specific checklist item based on VISIBLE geometry
3. Consider the material properties and production volume in your assessment
4. Factor in the casting type and process requirements

OUTPUT REQUIREMENTS - CRITICAL:
- Return ONLY a single JSON object
- NO markdown, NO code blocks, NO backticks, NO explanation text
- Use EXACTLY this format:

{{"result": "Yes", "reason": "Brief engineering justification considering {context['material']} and {context['volume']:,} parts", "confidence": "High"}}

RESULT VALUES:
- "Yes" = Drawing complies with this checklist item for the specified material and volume
- "No" = Drawing violates this checklist item (explain the violation)
- "Needs Review" = Cannot determine from visible geometry

CONFIDENCE VALUES:
- "High" = Clear evidence visible in drawing
- "Medium" = Partial evidence, some assumptions made
- "Low" = Limited visibility, uncertain assessment

RESPOND WITH JSON ONLY:"""

    try:
        # Prepare content for Gemini API (prompt + images)
        content = [prompt] + image_parts
        
        response = model.generate_content(content)
        response_text = response.text
        
        parsed = extract_json_from_response(response_text)
        
        if parsed:
            result = parsed.get("result", "Needs Review")
            reason = parsed.get("reason", "No reason provided")
            confidence = parsed.get("confidence", "Low")
            
            # Normalize result values
            result_upper = result.strip().lower()
            if result_upper in ["yes", "compliant", "pass", "true"]:
                result = "Yes"
            elif result_upper in ["no", "non-compliant", "fail", "false"]:
                result = "No"
            else:
                result = "Needs Review"
            
            return {
                "result": result,
                "reason": str(reason)[:500],
                "confidence": confidence
            }
        else:
            return {
                "result": "Needs Review",
                "reason": f"AI response parsing failed",
                "confidence": "Low"
            }
            
    except Exception as e:
        return {
            "result": "Needs Review",
            "reason": f"API error: {str(e)}",
            "confidence": "Low"
        }


def save_formatted_excel(checklist_rows, casting_context, output_dir):
    """Save results to formatted Excel file"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    material_short, volume_short = get_filename_components(casting_context)
    
    output_file = f"casting_analysis_{material_short}_{volume_short}parts_{timestamp}.xlsx"
    output_path = os.path.join(output_dir, output_file)

    # Create workbook with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Casting Analysis"
    
    # Add analysis parameters at the top
    ws.merge_cells('A1:H1')
    ws['A1'] = f"CASTING DESIGN ANALYSIS - {casting_context['casting_type']}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:H2')
    ws['A2'] = f"Material: {casting_context['material']} | Volume: {casting_context['volume']:,} parts | Process: {casting_context['process']}"
    ws['A2'].font = Font(size=10)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:H3')
    ws['A3'] = f"Tolerance: {casting_context['tolerance']} | Surface Finish: {casting_context['surface_finish']} | Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A3'].font = Font(size=9)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Style definitions
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers starting from row 5
    headers = ["Rule ID", "Rule Title", "Check ID", "Checklist Item", "Result (Yes/No)", "Notes / Observations", "Recommended Actions"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Data starting from row 6
    df = pd.DataFrame(checklist_rows)
    for r_idx, row in enumerate(df.values, 6):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            # Color code Result column (column 5)
            if c_idx == 5:
                if value == "Yes":
                    cell.fill = green_fill
                elif value == "No":
                    cell.fill = red_fill
    
    # Merge cells for Rule ID and Rule Title
    current_rule = None
    merge_start = 6
    
    for row_idx in range(6, len(df) + 7):
        if row_idx <= len(df) + 5:
            rule_id = ws.cell(row=row_idx, column=1).value
        else:
            rule_id = None
            
        if rule_id != current_rule:
            if current_rule is not None and row_idx - 1 > merge_start:
                ws.merge_cells(start_row=merge_start, start_column=1, end_row=row_idx-1, end_column=1)
                ws.cell(row=merge_start, column=1).alignment = Alignment(horizontal='center', vertical='center')
                
                ws.merge_cells(start_row=merge_start, start_column=2, end_row=row_idx-1, end_column=2)
                ws.cell(row=merge_start, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            merge_start = row_idx
            current_rule = rule_id
    
    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 60
    ws.column_dimensions['G'].width = 70
    
    wb.save(output_path)
    return output_file


def run_casting_analysis(excel_path, pdf_path, casting_context, output_dir):
    """
    Main analysis function - updated to match v1.py logic
    """
    try:
        print(f"DEBUG: Starting analysis with excel_path={excel_path}, pdf_path={pdf_path}")
        
        # Load rules from Excel
        rules_data = load_rules_from_excel(excel_path)
        print(f"DEBUG: Loaded {len(rules_data['rules'])} rules")
        
        # Convert PDF to images
        image_paths = pdf_to_images(pdf_path)
        print(f"DEBUG: Converted PDF to {len(image_paths)} images")
        
        # Pre-load images for Gemini API (if available)
        image_parts = []
        if GEMINI_AVAILABLE:
            for img in image_paths:
                # Use PIL to load image for Gemini API
                from PIL import Image
                image = Image.open(img)
                image_parts.append(image)
            print(f"DEBUG: Loaded {len(image_parts)} image parts for AI analysis")
        
        checklist_rows = []
        success_count = 0
        total_checks = 0
        
        for rule in rules_data["rules"]:
            print(f"DEBUG: Processing rule {rule['rule_id']} with {len(rule['checklist_items'])} items")
            for check_item in rule['checklist_items']:
                total_checks += 1
                print(f"DEBUG: Evaluating {check_item['check_id']}: {check_item['text'][:50]}...")
                
                # Evaluate checklist item
                result = evaluate_checklist_item(rule, check_item, image_parts, casting_context)
                print(f"DEBUG: Result for {check_item['check_id']}: {result['result']}")
                
                if "parsing failed" not in result["reason"].lower() and "api error" not in result["reason"].lower():
                    success_count += 1
                
                # Get recommended action for 'No' results
                recommended_action = get_recommended_action(rule, check_item, result['result'], casting_context)
                
                checklist_rows.append({
                    "Rule ID": rule["rule_id"],
                    "Rule Title": rule["title"],
                    "Check ID": check_item["check_id"],
                    "Checklist Item": check_item["text"],
                    "Result (Yes/No)": result["result"],
                    "Notes / Observations": result["reason"],
                    "Recommended Actions": recommended_action
                })
                
                # Rate limiting for API calls
                if GEMINI_AVAILABLE:
                    time.sleep(API_DELAY)
        
        print(f"DEBUG: Completed analysis. Total checks: {total_checks}, Success: {success_count}")
        
        # Save formatted Excel
        output_file = save_formatted_excel(checklist_rows, casting_context, output_dir)
        print(f"DEBUG: Saved Excel file: {output_file}")
        
        # Calculate summary statistics
        yes_count = sum(1 for row in checklist_rows if row["Result (Yes/No)"] == "Yes")
        no_count = sum(1 for row in checklist_rows if row["Result (Yes/No)"] == "No")
        review_count = sum(1 for row in checklist_rows if row["Result (Yes/No)"] == "Needs Review")
        
        print(f"DEBUG: Results - Yes: {yes_count}, No: {no_count}, Review: {review_count}")
        
        # Clean up temporary images
        for img_path in image_paths:
            try:
                os.remove(img_path)
            except:
                pass
        
        return {
            "total_checks": total_checks,
            "successful_evaluations": success_count,
            "results": {
                "compliant": yes_count,
                "non_compliant": no_count,
                "needs_review": review_count
            },
            "output_file": output_file,
            "details": checklist_rows[:10]  # First 10 items for preview
        }
        
    except Exception as e:
        print(f"DEBUG: Error in analysis: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            "error": str(e),
            "total_checks": 0,
            "successful_evaluations": 0,
            "results": {"compliant": 0, "non_compliant": 0, "needs_review": 0},
            "output_file": None,
            "details": []
        }
